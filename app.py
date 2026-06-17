# app.py
from __future__ import annotations
import io, re
from datetime import date, datetime, time, timedelta
from html import escape
import pandas as pd
import streamlit as st
from bs4 import BeautifulSoup

st.set_page_config(page_title="Avaya CMS: анализ смен", layout="wide")

AGENTS = {"7599449":"Абусаитов ДМ","7599415":"Аспенбетова АА","7599497":"Ахметзянова РР","7599437":"Воронцов ВВ","7599458":"Гайфуллина ДИ","7599473":"Галиева АР","7599413":"Гараев РР","7599498":"Заббаров АИ","7599405":"Зайнудтинова ЛС","7599411":"Ибрагимова ЛИ","7599403":"Минибаева АИ","7599408":"Сагетдинов МИ","7599478":"Хузахметов АР"}
SKILLS = {"1":"Надежность","3":"Качество э/э","9":"ЭЗС"}
ACCEPTED = {"ANS"}
MISSED = {"ABAN"}

def safe_download_button(*args, **kwargs):
    kwargs.setdefault("on_click", "ignore")
    try:
        return st.download_button(*args, **kwargs)
    except TypeError:
        kwargs.pop("on_click", None)
        return st.download_button(*args, **kwargs)

def decode(data: bytes) -> str:
    for enc in ("cp1251","windows-1251","utf-8","latin-1"):
        try: return data.decode(enc)
        except UnicodeDecodeError: pass
    return data.decode("utf-8", errors="replace")

def clean(x) -> str:
    return "" if x is None else str(x).replace("\xa0"," ").strip()

def find_col(df, names):
    lower = {str(c).lower().strip(): c for c in df.columns}
    for n in names:
        if n.lower().strip() in lower:
            return lower[n.lower().strip()]
    return None

@st.cache_data(show_spinner=False, max_entries=2)
def parse_file(data: bytes) -> pd.DataFrame:
    soup = BeautifulSoup(decode(data), "lxml")
    tables = soup.find_all("table")
    if not tables: return pd.DataFrame()
    def score(t):
        h = [clean(x.get_text(" ", strip=True)).lower() for x in t.find_all("th")]
        keys = ("размещение","split/skill","имена пользователей","время нач","date","disposition")
        return sum(3 for k in keys if any(k in x for x in h)) + len(h)
    table = max(tables, key=score)
    headers = [clean(x.get_text(" ", strip=True)) for x in table.find_all("th")]
    rows = []
    for tr in table.find_all("tr"):
        cells = tr.find_all("td")
        if cells:
            row = [clean(td.get_text(" ", strip=True)) for td in cells]
            rows.append((row + [""] * len(headers))[:len(headers)])
    return pd.DataFrame(rows, columns=headers)

def normalize(raw):
    if raw.empty: return pd.DataFrame(), ["HTML-таблица не найдена"]
    cd = find_col(raw, ["Дата","Date"]); ct = find_col(raw, ["Время нач.","Время нач","Time"]); cp = find_col(raw, ["Размещение","Disposition"])
    cs = find_col(raw, ["Split/Skill","Skill","Split"]); ca = find_col(raw, ["Имена пользователей","Agent","Пользователь","User"])
    if not cd or not ct or not cp: return pd.DataFrame(), ["Не найдены обязательные колонки: Дата, Время нач., Размещение"]
    df = raw.copy()
    dt_text = (df[cd].astype(str).str.strip()+" "+df[ct].astype(str).str.strip()).str.strip()
    df["dt_start"] = pd.to_datetime(dt_text, format="%d.%m.%Y %H:%M:%S", errors="coerce")
    if df["dt_start"].isna().all(): df["dt_start"] = pd.to_datetime(dt_text, dayfirst=True, errors="coerce")
    df = df.dropna(subset=["dt_start"]).reset_index(drop=True)
    if df.empty: return pd.DataFrame(), ["Дата/время не распознаны"]
    df["disposition"] = df[cp].astype(str).str.strip().str.upper()
    df["skill_code"] = ("" if cs is None else df[cs].astype(str).str.strip()).astype(str).str.extract(r"(\d+)", expand=False).fillna("")
    df["agent_code"] = "" if ca is None else df[ca].astype(str).str.strip()
    df["agent_code"] = df["agent_code"].replace({"nan":"","None":""})
    df["agent_name"] = df["agent_code"].map(AGENTS).fillna(df["agent_code"])
    df.loc[df["agent_code"].astype(str).str.strip()=="","agent_name"] = "Без оператора"
    df["call_date"] = df["dt_start"].dt.date
    df["slot"] = df["dt_start"].apply(lambda x: pd.Timestamp(x).replace(minute=0 if x.minute < 30 else 30, second=0, microsecond=0))
    return df, []

def shift_bounds(base: date, mode: str):
    if mode == "night":
        return datetime.combine(base, time(18,30)), datetime.combine(base + timedelta(days=1), time(6,30))
    return datetime.combine(base, time(6,30)), datetime.combine(base, time(18,30))

def available_dates(df, mode):
    dates = set(df["call_date"].dropna().tolist())
    if mode == "night": dates |= {d - timedelta(days=1) for d in dates}
    out = []
    for d in sorted(dates):
        s,e = shift_bounds(d, mode)
        if ((df["dt_start"] >= pd.Timestamp(s)) & (df["dt_start"] < pd.Timestamp(e))).any(): out.append(d)
    return out

def axis_slots(window):
    s,e = window
    start = pd.Timestamp(s).replace(minute=0, second=0, microsecond=0)
    end = pd.Timestamp(e)
    if end.minute or end.second: end = (end + pd.Timedelta(hours=1)).replace(minute=0, second=0, microsecond=0)
    return list(pd.date_range(start, end, freq="30min"))

def interval(slot):
    slot = pd.Timestamp(slot)
    return f"{slot:%H:%M}-{(slot + pd.Timedelta(minutes=30)):%H:%M}"

def title_for(window, ops):
    s,e = window
    prefix = f"{s:%d.%m}-{e:%d.%m.%Y}" if s.date()!=e.date() else f"{s:%d.%m.%Y}"
    names = [] if ops.empty else [x for x in dict.fromkeys(ops["Оператор"].astype(str).tolist()) if x and x!="Без оператора"]
    return f"{prefix} ({', '.join(names[:8]) if names else 'без операторов'})"

def build_metrics(df, watched, only_watched, conn_ok, window):
    ev = df.copy()
    ev["is_accepted"] = ev["disposition"].isin(ACCEPTED | ({"CONN"} if conn_ok else set()))
    ev["is_missed"] = ev["disposition"].isin(MISSED)
    if only_watched and watched: ev["is_missed"] = ev["is_missed"] & ev["skill_code"].isin(watched)
    ev["is_missed_no_agent"] = ev["is_missed"] & (ev["agent_code"].astype(str).str.strip()=="")
    ev["is_missed_with_agent"] = ev["is_missed"] & (ev["agent_code"].astype(str).str.strip()!="")
    ev["skill_name"] = ev["skill_code"].map(SKILLS).fillna(ev["skill_code"].replace("", "Без тематики"))
    ev.loc[ev["skill_name"].astype(str).str.strip()=="","skill_name"] = "Без тематики"

    dyn = pd.DataFrame({"slot": axis_slots(window)})
    g = ev.groupby("slot").agg(Принято=("is_accepted","sum"), Пропущено=("is_missed","sum"), **{"Пропущено без оператора":("is_missed_no_agent","sum")}, Всего=("disposition","size")).reset_index()
    dyn = dyn.merge(g, on="slot", how="left").fillna(0)
    for c in ("Принято","Пропущено","Пропущено без оператора","Всего"): dyn[c] = dyn[c].astype(int)
    dyn["Порядок"] = range(1, len(dyn)+1)
    dyn["Ось"] = dyn.apply(lambda r: f"{int(r['Порядок']):02d} | {pd.Timestamp(r['slot']):%H:%M}", axis=1)
    dyn["Время"] = dyn["slot"].dt.strftime("%H:%M")
    dyn["Интервал"] = dyn["slot"].apply(interval)
    dyn = dyn[["Порядок","Ось","Время","Интервал","Принято","Пропущено","Пропущено без оператора","Всего","slot"]]

    agent_ev = ev[ev["agent_code"].astype(str).str.strip()!=""]
    if agent_ev.empty:
        ops = pd.DataFrame(columns=["Оператор","Принято","Пропущено","Всего","% пропущенных"])
    else:
        ops = agent_ev.groupby("agent_name").agg(Принято=("is_accepted","sum"), Пропущено=("is_missed","sum"), Всего=("disposition","size")).reset_index().rename(columns={"agent_name":"Оператор"})
        denom = (ops["Принято"]+ops["Пропущено"]).replace(0, pd.NA)
        ops["% пропущенных"] = (100*ops["Пропущено"]/denom).fillna(0).round(2)
        ops = ops.sort_values(["Пропущено","Принято"], ascending=[False,False])

    ttl = title_for(window, ops)
    shift_resp = f"Смена ({ttl.split('(',1)[1].rstrip(')')})"
    missed = ev[ev["is_missed"]].copy().sort_values("dt_start").reset_index(drop=True)
    if missed.empty:
        miss_det = pd.DataFrame(columns=["Порядок","Дата и время","Оператор Avaya","Ответственность","Тематика"])
    else:
        missed["Порядок"] = range(1, len(missed)+1)
        missed["Дата и время"] = missed["dt_start"].dt.strftime("%d.%m.%Y %H:%M:%S")
        missed["Оператор Avaya"] = missed["agent_name"]
        missed["Ответственность"] = missed.apply(lambda r: r["agent_name"] if str(r["agent_code"]).strip() else shift_resp, axis=1)
        missed["Тематика"] = missed["skill_name"]
        miss_det = missed[["Порядок","Дата и время","Оператор Avaya","Ответственность","Тематика"]]

    acc_map, miss_map, no_map = {}, {}, {}
    for slot, part in ev[ev["is_accepted"] & (ev["agent_code"].astype(str).str.strip()!="")].groupby("slot"):
        acc_map[pd.Timestamp(slot)] = "; ".join(f"{k}: {v}" for k,v in part.groupby("agent_name").size().sort_values(ascending=False).items())
    for slot, part in ev[ev["is_missed"]].groupby("slot"):
        resp = part.apply(lambda r: r["agent_name"] if str(r["agent_code"]).strip() else shift_resp, axis=1)
        miss_map[pd.Timestamp(slot)] = "; ".join(f"{k}: {v}" for k,v in resp.value_counts().items())
        no_map[pd.Timestamp(slot)] = int(part["is_missed_no_agent"].sum())

    scheme = dyn.copy()
    scheme["Кто принял"] = scheme["slot"].map(lambda x: acc_map.get(pd.Timestamp(x),""))
    scheme["Кто пропустил / ответственность"] = scheme["slot"].map(lambda x: miss_map.get(pd.Timestamp(x),""))
    scheme["Без оператора"] = scheme["slot"].map(lambda x: no_map.get(pd.Timestamp(x),0)).astype(int)
    scheme = scheme[["Порядок","Ось","Время","Принято","Кто принял","Пропущено","Кто пропустил / ответственность","Без оператора","Всего"]]

    total = int(ev["is_missed"].sum()); no_agent = int(ev["is_missed_no_agent"].sum()); in_ops = int(ops["Пропущено"].sum()) if not ops.empty else 0
    dist = pd.DataFrame([{ "Показатель":"Всего пропущено","Значение":total},{"Показатель":"Пропущено с кодом оператора","Значение":int(ev["is_missed_with_agent"].sum())},{"Показатель":"Пропущено без кода оператора","Значение":no_agent},{"Показатель":"Пропущено в таблице операторов","Значение":in_ops},{"Показатель":"Ответственность смены за пропущенные без оператора","Значение":no_agent},{"Показатель":"Контрольное расхождение","Значение":total-in_ops-no_agent}])
    skills = ev.groupby("skill_name").agg(Принято=("is_accepted","sum"), Пропущено=("is_missed","sum"), Всего=("disposition","size")).reset_index().rename(columns={"skill_name":"Тематика"}).sort_values(["Пропущено","Принято"], ascending=[False,False])
    peaks = dyn.drop(columns=["slot"]).sort_values(["Всего","Пропущено"], ascending=[False,False])
    anomalies = dyn[(dyn["Пропущено"]>0) & ((dyn["Принято"]==0) | (dyn["Пропущено без оператора"]>0))].drop(columns=["slot"]).copy()
    accepted = int(ev["is_accepted"].sum()); denom = accepted + total
    kpi = pd.DataFrame([{ "Принято":accepted,"Пропущено":total,"Пропущено без оператора":no_agent,"% пропущенных":round(100*total/denom,2) if denom else 0,"Всего событий":len(ev)}])
    return {"kpi":kpi,"distribution":dist,"dynamics":dyn.drop(columns=["slot"]),"scheme":scheme,"missed_details":miss_det,"operators":ops,"skills":skills,"peaks":peaks,"anomalies":anomalies,"events":ev}, ttl

def show(df, rows=300):
    if df is None or df.empty: st.caption("Нет данных")
    else: st.dataframe(df.head(rows), use_container_width=True)

def to_html(df, limit=700):
    return "<p>Нет данных</p>" if df is None or df.empty else df.head(limit).to_html(index=False, border=1, escape=True)

def make_word(m, summary, title):
    html = f"""<html><head><meta charset="utf-8"><style>@page {{size:A4 landscape;margin:10mm}} body{{font-family:Arial;font-size:9pt}} table{{border-collapse:collapse;width:100%;margin-bottom:12px}} th,td{{border:1px solid #999;padding:4px;vertical-align:top}} th{{background:#f2f2f2}}</style></head><body><h1>Анализ смены Avaya CMS</h1><h2>{escape(title)}</h2><h2>Вывод</h2><ul>{''.join(f'<li>{escape(str(x))}</li>' for x in summary)}</ul><h2>KPI</h2>{to_html(m['kpi'])}<h2>Проверка распределения пропущенных</h2>{to_html(m['distribution'])}<h2>Динамика по порядку смены</h2><p>Ось: 18:00 -> 07:00. 00:00 находится в середине.</p>{to_html(m['dynamics'],1000)}<h2>Схема смены</h2>{to_html(m['scheme'],1000)}<h2>Кто пропустил</h2>{to_html(m['missed_details'],1000)}<h2>Операторы</h2>{to_html(m['operators'])}<h2>Тематики</h2>{to_html(m['skills'])}<h2>Пики</h2>{to_html(m['peaks'])}<h2>Аномалии</h2>{to_html(m['anomalies'])}</body></html>"""
    return html.encode("utf-8")

def make_excel(m, summary, title):
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as w:
        pd.DataFrame({"Вывод": summary}).to_excel(w, index=False, sheet_name="Итог")
        for k,s in {"distribution":"Проверка","dynamics":"Динамика","scheme":"Схема смены","missed_details":"Кто пропустил","operators":"Операторы","skills":"Тематики","peaks":"Пики","anomalies":"Аномалии"}.items(): m[k].to_excel(w, index=False, sheet_name=s)
        wb = w.book; ws = wb["Динамика"]; chart_ws = wb.create_sheet("График",1)
        chart_ws["A1"] = "Динамика по порядку смены"; chart_ws["A2"] = "Ось: 18:00 -> 07:00. 00:00 находится в середине."
        if ws.max_row >= 2:
            chart = BarChart(); chart.title = "Принято / пропущено по получасам"
            data = Reference(ws, min_col=5, max_col=7, min_row=1, max_row=ws.max_row)
            cats = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)
            chart.add_data(data, titles_from_data=True); chart.set_categories(cats); chart.width = 38; chart.height = 16
            chart_ws.add_chart(chart, "A4")
        for ws in wb.worksheets:
            ws.page_setup.orientation = "landscape"; ws.page_setup.fitToWidth = 1; ws.freeze_panes = "A2"
            for cell in ws[1]:
                cell.font = Font(bold=True); cell.fill = PatternFill("solid", fgColor="EAF2F8"); cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for i in range(1, min(ws.max_column, 12)+1): ws.column_dimensions[get_column_letter(i)].width = 24
    return bio.getvalue()

for k,v in {"exports_ready":False,"excel":None,"word":None}.items(): st.session_state.setdefault(k,v)

st.title("Avaya CMS — анализ смен")
with st.sidebar:
    uploaded = st.file_uploader("HTML отчет Avaya", type=["html","htm"])
    mode = st.radio("Смена", ["night","day"], format_func=lambda x: "Ночная (18:30-06:30)" if x=="night" else "Дневная (06:30-18:30)")
    conn_ok = st.checkbox("Считать CONN как принято", value=False)
    only_watched = st.checkbox("Пропущенные считать только по выбранным skills", value=(mode=="night"))
    watched = st.multiselect("Skills", list(SKILLS.keys()), default=list(SKILLS.keys()))
    show_chart = st.checkbox("Показывать график", value=True)

if uploaded is None:
    st.info("Загрузи HTML отчет Avaya слева."); st.stop()

raw = parse_file(uploaded.getvalue()); df, problems = normalize(raw)
if problems: st.error("; ".join(problems)); st.stop()
dates = available_dates(df, mode)
if not dates: st.warning("В выбранной смене нет событий."); st.stop()
day = st.selectbox("Дата смены", dates, format_func=lambda d: d.strftime("%d.%m.%Y"))
window = shift_bounds(day, mode); df_shift = df[(df["dt_start"]>=pd.Timestamp(window[0])) & (df["dt_start"]<pd.Timestamp(window[1]))].copy()
metrics, title = build_metrics(df_shift, watched, only_watched, conn_ok, window)
kpi = metrics["kpi"].iloc[0].to_dict(); diff = int(metrics["distribution"].loc[metrics["distribution"]["Показатель"]=="Контрольное расхождение","Значение"].iloc[0])
summary = [f"Окно анализа: {window[0]:%d.%m.%Y %H:%M} - {window[1]:%d.%m.%Y %H:%M}.", f"Итог: принято {int(kpi['Принято'])}, пропущено {int(kpi['Пропущено'])}, без кода оператора {int(kpi['Пропущено без оператора'])}.", "Контроль распределения пропущенных: расхождение 0." if diff==0 else f"Контроль распределения пропущенных: расхождение {diff}.", "Динамика построена по порядку смены: 18:00 -> 07:00, 00:00 в середине."]
st.subheader(title)
for x in summary: st.write("- " + x)
c1,c2,c3,c4 = st.columns(4); c1.metric("Принято", int(kpi["Принято"])); c2.metric("Пропущено", int(kpi["Пропущено"])); c3.metric("Без оператора", int(kpi["Пропущено без оператора"])); c4.metric("% пропущенных", f"{kpi['% пропущенных']}%")
st.markdown("### Проверка распределения пропущенных"); show(metrics["distribution"])
st.markdown("### Динамика по порядку смены"); st.caption("Ось построена строго по порядку смены: 18:00 -> 07:00. 00:00 находится в середине."); show(metrics["dynamics"], rows=1000)
if show_chart: st.bar_chart(metrics["dynamics"].set_index("Порядок")[["Принято","Пропущено","Пропущено без оператора"]]); st.caption("На графике X = Порядок смены, расшифровка порядка выше.")
st.markdown("### Схема смены"); show(metrics["scheme"], rows=1000)
st.markdown("### Кто и во сколько пропустил"); show(metrics["missed_details"], rows=1000)
left,right = st.columns(2)
with left: st.markdown("### Операторы"); show(metrics["operators"])
with right: st.markdown("### Тематики"); show(metrics["skills"])
with st.expander("Пики и аномалии"): show(metrics["peaks"]); show(metrics["anomalies"])
st.markdown("### Экспорт")
if st.button("Подготовить Word и Excel", use_container_width=True):
    try:
        st.session_state.excel = make_excel(metrics, summary, title); st.session_state.word = make_word(metrics, summary, title); st.session_state.exports_ready = True; st.success("Файлы готовы.")
    except Exception as exc:
        st.session_state.exports_ready = False; st.exception(exc)
if st.session_state.exports_ready:
    base = re.sub(r'[\\/:*?"<>|]+', "-", title)[:160]; a,b = st.columns(2)
    with a: safe_download_button("Скачать Excel", st.session_state.excel, f"{base} avaya_report.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
    with b: safe_download_button("Скачать Word", st.session_state.word, f"{base} avaya_report.doc", "application/msword", use_container_width=True)
